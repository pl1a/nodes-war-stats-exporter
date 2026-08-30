"""
Copyright 2026 pl1a

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated documentation files (the “Software”), to deal in the Software without restriction, including without limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED “AS IS”, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

from collections import defaultdict
import re
import requests
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTALS_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTALS_FONT = Font(bold=True)
ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # light blue
THIN_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
)


def fetch_player_nations(url: str) -> dict[str, str]:
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    name_to_nation: dict[str, str] = {}
    for resident in data.get("residents", {}).values():
        name = resident.get("name")
        nation = resident.get("nation")
        if name:
            name_to_nation[name] = nation if nation else "No Nation"
    return name_to_nation


def style_header_row(worksheet, num_cols: int, header_row: int = 1, start_col: int = 1):
    for col_idx in range(start_col, start_col + num_cols):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")


def style_data_rows(
    worksheet,
    num_rows: int,
    num_cols: int,
    header_row: int = 1,
    start_col: int = 1,
):
    for r in range(num_rows):
        excel_row = header_row + 1 + r
        fill = ROW_FILL_EVEN if r % 2 == 0 else ROW_FILL_ODD
        for c in range(num_cols):
            cell = worksheet.cell(row=excel_row, column=start_col + c)
            cell.fill = fill
            cell.border = THIN_BORDER


def add_totals_row(
    worksheet,
    df: pd.DataFrame,
    has_nation_col: bool = False,
    start_col: int = 1,
    header_row: int = 1,
):
    if df.empty:
        return

    total_kills = int(df["Kills"].sum())
    total_deaths = int(df["Deaths"].sum())
    total_kda = round(total_kills / max(total_deaths, 1), 2)

    totals_row = header_row + len(df) + 1

    if has_nation_col:
        values = ["TOTAL", "", total_kills, total_deaths, total_kda]
    else:
        values = ["TOTAL", total_kills, total_deaths, total_kda]

    for i, value in enumerate(values):
        cell = worksheet.cell(row=totals_row, column=start_col + i, value=value)
        cell.font = TOTALS_FONT
        cell.fill = TOTALS_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if i > 0 else "left")

    last_data_row = header_row + len(df)
    first_col = get_column_letter(start_col)
    last_col = get_column_letter(start_col + len(df.columns) - 1)
    worksheet.auto_filter.ref = f"{first_col}{header_row}:{last_col}{last_data_row}"


def set_column_widths(worksheet, df: pd.DataFrame, start_col: int = 1):
    for i, col_name in enumerate(df.columns):
        col_idx = start_col + i
        max_len = max(
            df[col_name].astype(str).map(len).max() if not df.empty else 0,
            len(str(col_name)),
            6,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def parse_chat_logs(
    log_file_path: str,
    output_excel_path: str,
    towns_url: str,
):
    pattern = re.compile(
        r"\[CHAT\]\s+(?P<victim>[^\s]+)\s+was\s+.+?\s+by\s+(?P<killer>[^\s]+)"
    )
    stats = defaultdict(lambda: {"Kills": 0, "Deaths": 0})

    with open(log_file_path, "r", encoding="utf-8") as file:
        for line in file:
            match = pattern.search(line)
            if match:
                victim = match.group("victim")
                killer = match.group("killer")
                stats[victim]["Deaths"] += 1
                stats[killer]["Kills"] += 1

    print(f"Downloading player/nation data from {towns_url} ...")
    name_to_nation = fetch_player_nations(towns_url)
    print(f"Loaded {len(name_to_nation)} players from towns.json")

    data = []
    for player, counts in stats.items():
        kills = counts["Kills"]
        deaths = counts["Deaths"]
        kd_ratio = round(kills / max(deaths, 1), 2)
        nation = name_to_nation.get(player, "Unknown")
        data.append(
            {
                "Player": player,
                "Nation": nation,
                "Kills": kills,
                "Deaths": deaths,
                "KD/A": kd_ratio,
            }
        )

    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by="Kills", ascending=False).reset_index(drop=True)

    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Player Stats", index=False)
        ws = writer.sheets["Player Stats"]
        if not df.empty:
            style_header_row(ws, num_cols=len(df.columns), header_row=1, start_col=1)
            style_data_rows(ws, num_rows=len(df), num_cols=len(df.columns), header_row=1, start_col=1)
            set_column_widths(ws, df, start_col=1)
            add_totals_row(ws, df, has_nation_col=True, start_col=1, header_row=1)

            nation_summary = (
                df.groupby("Nation", as_index=False)
                .agg(
                    Rally=("Player", "count"),
                    Kills=("Kills", "sum"),
                    Deaths=("Deaths", "sum"),
                )
            )
            nation_summary["KD/A"] = (
                nation_summary["Kills"] / nation_summary["Deaths"].clip(lower=1)
            ).round(2)
            nation_summary["_sort"] = nation_summary["Nation"].map(
                lambda x: (x == "No Nation", x == "Unknown", 0)
            )
            nation_summary = (
                nation_summary
                .sort_values(by=["_sort", "Rally"], ascending=[True, False])
                .drop(columns=["_sort"])
                .reset_index(drop=True)
            )

            start_col = 7
            headers = ["Nation", "Rally", "Kills", "Deaths", "KD/A"]

            header_row = 1
            for i, h in enumerate(headers):
                cell = ws.cell(row=header_row, column=start_col + i, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            for i, row in nation_summary.iterrows():
                r = header_row + 1 + i
                fill = ROW_FILL_EVEN if i % 2 == 0 else ROW_FILL_ODD
                values = [
                    row["Nation"],
                    int(row["Rally"]),
                    int(row["Kills"]),
                    int(row["Deaths"]),
                    row["KD/A"],
                ]
                for j, val in enumerate(values):
                    cell = ws.cell(row=r, column=start_col + j, value=val)
                    cell.fill = fill
                    cell.border = THIN_BORDER
                    if j > 0:
                        cell.alignment = Alignment(horizontal="center")

            totals_row = header_row + 1 + len(nation_summary)
            total_rally = int(nation_summary["Rally"].sum())
            total_kills = int(nation_summary["Kills"].sum())
            total_deaths = int(nation_summary["Deaths"].sum())
            total_kda = round(total_kills / max(total_deaths, 1), 2)
            totals_values = ["TOTAL", total_rally, total_kills, total_deaths, total_kda]
            for j, val in enumerate(totals_values):
                cell = ws.cell(row=totals_row, column=start_col + j, value=val)
                cell.font = TOTALS_FONT
                cell.fill = TOTALS_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center" if j > 0 else "left")

            for i, col_name in enumerate(headers):
                col_letter = get_column_letter(start_col + i)
                needed = max(len(str(col_name)) + 2, 12)
                if i == 0 and not nation_summary.empty:
                    needed = max(
                        needed,
                        int(nation_summary["Nation"].astype(str).map(len).max()) + 2,
                    )
                ws.column_dimensions[col_letter].width = min(needed, 40)

        if not df.empty:
            nations = sorted(
                df["Nation"].unique(),
                key=lambda x: (x == "No Nation", x == "Unknown", x),
            )

            for nation in nations:
                nation_df = (
                    df[df["Nation"] == nation]
                    .drop(columns=["Nation"])
                    .sort_values(by="Kills", ascending=False)
                    .reset_index(drop=True)
                )

                sheet_name = re.sub(r'[\\/*?:\[\]]', "_", str(nation))[:31]
                if not sheet_name:
                    sheet_name = "Unknown"

                nation_df.to_excel(writer, sheet_name=sheet_name, index=False)
                nws = writer.sheets[sheet_name]
                if not nation_df.empty:
                    style_header_row(nws, num_cols=len(nation_df.columns), header_row=1, start_col=1)
                    style_data_rows(nws, num_rows=len(nation_df), num_cols=len(nation_df.columns), header_row=1, start_col=1)
                    set_column_widths(nws, nation_df, start_col=1)
                    add_totals_row(nws, nation_df, has_nation_col=False, start_col=1, header_row=1)

    print(
        f"Player statistics successfully exported to "
        f"{output_excel_path}"
    )
    if not df.empty:
        print(f"Sheets created: Player Stats + {len(nations)} nation sheets")

